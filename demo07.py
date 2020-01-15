import matplotlib.pyplot as plt
import matplotlib
import smtplib
import openpyxl as op
from openpyxl.drawing.image import Image
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.header import Header

# 准备数据
x_list = []
y_list = []
# 打开Excel文件
wbr = op.load_workbook('sp500.xlsx')
sheet = wbr.get_sheet_by_name(wbr.sheetnames[0])
a = sheet['A']
b = sheet['B']
for j in a:
    x_list.append(j.value)
for i in b:
    y_list.append(i.value)
# 设置中文字体正常显示
matplotlib.rcParams['font.sans-serif'] = ['SimHei']
plt.plot(x_list, y_list, label='走势图')
plt.xlabel('时间')
plt.ylabel('指数')
plt.title('标准普尔500指数')
plt.legend()
plt.savefig('500.png')
img = Image('500.png')
wbr.create_sheet('走势图')
new_sheet = wbr.get_sheet_by_name(wbr.sheetnames[1])
new_sheet.add_image(img)
wbr.save('sp500.xlsx')
# 使用邮件发送文件
smtpserver = 'smtp.qq.com'
username = 'xxx@qq.com'
password='授权码'
sender='xxx@qq.com'
receiver=['xxx@sina.com']
subject = 'python发送报表'
subject=Header(subject, 'utf-8').encode()
#构造邮件对象MIMEMultipart对象
#下面的主题，发件人，收件人，日期是显示在邮件页面上的。
msg = MIMEMultipart('mixed') 
msg['Subject'] = subject
msg['From'] = sender
#收件人为多个收件人,通过join将列表转换为以;为间隔的字符串
msg['To'] = ";".join(receiver) 
#构造文字内容  
text = "你好，这是由python发送的邮件！报表见附件"    
text_plain = MIMEText(text,'plain', 'utf-8')    
msg.attach(text_plain)    
#构造附件
sendfile=open('sp500.xlsx','rb').read()
text_att = MIMEText(sendfile, 'base64', 'utf-8') 
text_att["Content-Type"] = 'application/octet-stream'  
#以下附件可以重命名
text_att.add_header('Content-Disposition', 'attachment', filename='sp500.xlsx')
msg.attach(text_att)
try:
    server = smtplib.SMTP_SSL(smtpserver, 465)
    server.ehlo()
    #登录你的账号
    server.login(username, password)
    #发送邮件
    server.sendmail(sender, receiver, msg.as_string())
    print("邮件发送成功")
    server.quit()  # 关闭连接
except Exception as e:
    print("邮件发送失败"+e)